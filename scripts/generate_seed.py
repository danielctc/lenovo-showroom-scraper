#!/usr/bin/env python3
"""Parse a Lenovo TopSeller xlsx into a D1 seed.sql file.

Reads the "All Part Numbers" canonical sheet (PRD §9) and emits INSERT
statements for the products, product_specs (stub: processor/memory/storage
+ OS/weight/warranty from the flat table) and product_images (one
placeholder row per product with r2_key='external' and a best-effort
PSREF image URL).

Usage:
  python3 generate_seed.py --xlsx ../fixtures/TS_Price_List_T2_April_2026.xlsx \\
                           --out /tmp/seed/seed.sql

The output file is meant to be applied against a FRESH D1 database via:
  wrangler d1 execute SHOWROOM_DB --remote --file=/tmp/seed/seed.sql

For subsequent quarterly refreshes against an existing DB, use
`scrape_all.py` and `apply_to_d1.py` instead — those do targeted
UPDATE/DELETE+INSERT and won't duplicate rows.
"""
from __future__ import annotations
import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# Xlsx brand column → showroom category tile
BRAND_TO_CATEGORY = {
    "ThinkCentre": "Desktops",
    "Lenovo": "Laptops",
    "ThinkPad": "ThinkPad",
    "ThinkBook": "ThinkPad",
    "Lenovo Tablets": "Tablets",
    "ThinkSmart": "Smart Office",
    "ThinkVision": "Monitors",
    "ThinkStation": "Workstations",
    "Legion": "Laptops",
}

# Canonical sheet column indexes (0-indexed, All Part Numbers sheet, no header)
COL_PART_NUMBER = 1
COL_BRAND = 3
COL_FAMILY = 5
COL_NAME = 9
COL_ERP_UK = 10
COL_ERP_IE = 11
COL_WEIGHT = 13
COL_WARRANTY = 14
COL_EAN = 15
COL_OS = 16
COL_PROCESSOR = 18
COL_MEMORY = 19
COL_STORAGE = 20


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    # Some cells use `^|^` as a pseudo-separator
    if "^|^" in s:
        s = s.split("^|^")[0].strip()
    return s or None


def clean_ean(v):
    s = clean(v)
    if s and " / " in s:
        s = s.split(" / ")[0].strip()
    return s if s and s.lower() != "none" else None


def price_num(s) -> str:
    if not s:
        return "NULL"
    m = re.match(r"([£€])([\d,]+)", str(s))
    if not m:
        return "NULL"
    v = int(m.group(2).replace(",", ""))
    # Guard against pence-in-pounds errors (seen on the Legion row in T2 2026)
    if v > 50_000:
        v = v / 100
    return str(v)


def sql(v) -> str:
    if v is None or v == "":
        return "NULL"
    s = str(v).strip()
    if "^|^" in s:
        s = s.split("^|^")[0].strip()
    s = s.replace("'", "''")
    return "'" + s + "'"


def image_url(brand: str, family_raw: str) -> str | None:
    if not brand or not family_raw:
        return None
    brand_slug = brand.replace(" ", "_")
    fam = re.sub(r"\s*\(.+?\)\s*", "", family_raw).strip()
    fam = fam.replace(" Monitor", "").replace(" ", "_")
    return (
        f"https://psrefstuff.lenovo.com/syspool//Sys/Image/{brand_slug}/{fam}"
        f"/Compressedimage/{fam}_CT1_01.png"
    )


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path)
    ap.add_argument("--sheet", default="All Part Numbers", help="Canonical sheet name")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"error: sheet {args.sheet!r} not in workbook (have {wb.sheetnames})", file=sys.stderr)
        return 1
    ws = wb[args.sheet]

    seen: set[str] = set()
    products_rows: list[str] = []
    specs_rows: list[str] = []
    images_rows: list[str] = []

    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 21:
            continue
        pn = r[COL_PART_NUMBER]
        brand = r[COL_BRAND]
        family_raw = r[COL_FAMILY]
        name = r[COL_NAME]
        erp_uk = r[COL_ERP_UK]
        erp_ie = r[COL_ERP_IE]
        if not pn or pn in seen or not name or not erp_uk:
            continue
        seen.add(pn)
        weight = clean(r[COL_WEIGHT])
        warranty = clean(r[COL_WARRANTY])
        ean = clean_ean(r[COL_EAN])
        os_ = clean(r[COL_OS])
        proc = clean(r[COL_PROCESSOR])
        mem = clean(r[COL_MEMORY])
        stor = clean(r[COL_STORAGE])
        category = BRAND_TO_CATEGORY.get(brand)
        if not category:
            continue  # skip Software / Accessories / Education (no showroom home)
        euk = price_num(erp_uk)
        eie = price_num(erp_ie)
        luk = f"{float(euk) * 1.15:.2f}" if euk != "NULL" else "NULL"
        lie = f"{float(eie) * 1.15:.2f}" if eie != "NULL" else "NULL"
        img = image_url(brand, family_raw)
        now = "datetime('now')"
        products_rows.append(
            "INSERT INTO products (part_number,name,series,family,category,"
            "erp_price_uk,list_price_uk,erp_price_ie,list_price_ie,warranty,weight,ean,"
            "psref_url,is_active,created_at,updated_at) VALUES ("
            f"{sql(pn)},{sql(name)},{sql(brand)},{sql(family_raw)},{sql(category)},"
            f"{euk},{luk},{eie},{lie},{sql(warranty)},{sql(weight)},{sql(ean)},"
            f"'https://psref.lenovo.com/Detail/?M={pn}',1,{now},{now});"
        )
        if proc:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Performance','Processor',{sql(proc)},0);")
        if mem:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Performance','Memory',{sql(mem)},1);")
        if stor:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Performance','Storage',{sql(stor)},2);")
        if os_:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Power & OS','Operating System',{sql(os_)},0);")
        if weight:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Power & OS','Weight',{sql(weight)},1);")
        if warranty:
            specs_rows.append(f"INSERT INTO product_specs (part_number,category,label,value,sort_order) VALUES ({sql(pn)},'Power & OS','Warranty',{sql(warranty)},2);")
        if img:
            images_rows.append(f"INSERT INTO product_images (part_number,source_url,r2_key,sort_order) VALUES ({sql(pn)},{sql(img)},'external',0);")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w") as fh:
        fh.write("-- Seed generated from {}\n".format(args.xlsx.name))
        fh.write("-- NOTE: no BEGIN/COMMIT — D1 rejects explicit transactions, it wraps each exec in a batch itself.\n")
        for row in products_rows:
            fh.write(row + "\n")
        for row in specs_rows:
            fh.write(row + "\n")
        for row in images_rows:
            fh.write(row + "\n")

    print(f"[info] wrote {args.out}")
    print(f"[info] rows: products={len(products_rows)} specs={len(specs_rows)} images={len(images_rows)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
