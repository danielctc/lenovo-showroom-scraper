"""Parse the TopSeller spreadsheet.

Canonical source = the "All Part Numbers" sheet (517 rows, 48 cols) per PRD §9.
Per-sheet column mappings live in config/sheet_mapping.json for drift resilience.
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import json

from openpyxl import load_workbook


CANONICAL_SHEET = "All Part Numbers"


@dataclass
class CommercialRow:
    part_number: str
    name: str
    series: str | None = None
    family: str | None = None
    category: str | None = None
    erp_price_uk: float | None = None
    list_price_uk: float | None = None
    erp_price_ie: float | None = None
    list_price_ie: float | None = None
    trade_in_uk: str | None = None
    trade_in_ie: str | None = None
    leap: float | None = None
    warranty: str | None = None
    weight: str | None = None
    ean: str | None = None
    psref_url: str | None = None


def load_mapping(config_path: Path) -> dict:
    return json.loads(config_path.read_text())


def parse(spreadsheet: Path, mapping_path: Path | None = None) -> list[CommercialRow]:
    """Read the canonical "All Part Numbers" sheet and return CommercialRow objects."""
    mapping = load_mapping(mapping_path) if mapping_path else {}
    wb = load_workbook(spreadsheet, read_only=True, data_only=True)
    if CANONICAL_SHEET not in wb.sheetnames:
        raise ValueError(f"spreadsheet missing required sheet: {CANONICAL_SHEET!r}")
    ws = wb[CANONICAL_SHEET]
    # TODO: build header index, iterate rows, coerce types per mapping.
    _ = mapping, ws
    return []
